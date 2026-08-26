from __future__ import annotations

import difflib
import html
import os
import re
import threading
import time
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Protocol
from urllib.parse import urlparse


@dataclass(frozen=True)
class ProductVariant:
    model: str
    memory: str
    color: str
    price_uzs: Decimal
    product_id: int | None = None
    url: str | None = None
    warranty_months: int | None = None


@dataclass(frozen=True)
class ProductMatch:
    status: str
    models: tuple[str, ...] = ()
    variants: tuple[ProductVariant, ...] = ()
    model_urls: tuple[tuple[str, str], ...] = ()
    requested_memory: str | None = None
    requested_color: str | None = None
    unmatched_filters: tuple[str, ...] = ()
    # The selected rows may be filtered by a customer-supplied memory/color.
    # Keep the approved model's complete rows in the same lookup so the wizard
    # can show every real choice without a second source request.
    all_variants: tuple[ProductVariant, ...] = ()

    def url_for(self, model: str) -> str | None:
        return next((url for name, url in self.model_urls if name == model), None)

    @property
    def filters_matched(self) -> bool:
        return not self.unmatched_filters


class ProductRepository(Protocol):
    def search(
        self,
        query: str,
        memory: str | None = None,
        color: str | None = None,
    ) -> ProductMatch: ...


ALIASES = {
    "айфон": "iphone",
    "айфона": "iphone",
    "айфончик": "iphone",
    "ayfon": "iphone",
    "aifon": "iphone",
    "эпл": "apple",
    "эппл": "apple",
    "самсунг": "samsung",
    "samsumg": "samsung",
    "samnsung": "samsung",
    "samsun": "samsung",
    "samsng": "samsung",
    "samsug": "samsung",
    "samsungg": "samsung",
    "iphnoe": "iphone",
    "iphnone": "iphone",
    "ipone": "iphone",
    "ifon": "iphone",
    "айфоон": "iphone",
    "редми": "redmi",
    "сяоми": "xiaomi",
    "ксяоми": "xiaomi",
    "хуавей": "huawei",
    "поко": "poco",
    "хонор": "honor",
    "про": "pro",
    "промакс": "pro max",
    "promax": "pro max",
    "макс": "max",
    "максс": "max",
    "мах": "max",
    "ультра": "ultra",
    "ультраа": "ultra",
    "ulrta": "ultra",
    "ulttra": "ultra",
    "плюс": "plus",
    "черный": "black", "черная": "black", "черного": "black", "черном": "black", "qora": "black", "qorasi": "black", "қора": "black",
    "белый": "white", "белая": "white", "белого": "white", "белом": "white", "oq": "white", "оқ": "white",
    "синий": "blue", "синяя": "blue", "синем": "blue", "голубой": "blue", "kok": "blue", "ko'k": "blue", "кўк": "blue",
    "зеленый": "green", "зеленая": "green", "зеленом": "green", "yashil": "green", "яшил": "green",
    "красный": "red", "красная": "red", "красном": "red", "qizil": "red", "қизил": "red",
    "серый": "gray", "серая": "gray", "сером": "gray", "серебристый": "silver",
    "золотой": "gold", "золотая": "gold", "розовый": "pink", "розовая": "pink",
    "фиолетовый": "purple", "фиолетовая": "purple", "фиолетового": "purple",
    "binafsha": "purple", "бинафша": "purple", "moviy": "blue", "мовий": "blue",
    "натуральный": "natural", "натуральная": "natural", "титановый": "titanium",
    "титановая": "titanium", "титан": "titanium", "мокко": "mocha",
}

TRUSTED_POST_HOSTS = {"t.me", "telegram.me", "www.t.me", "www.telegram.me"}
SEARCH_FILLERS = {
    "price", "narx", "narxi", "narxini", "narxlar", "qancha", "qanchadan", "necha", "цена", "цену", "цены",
    "сколько", "стоит", "почем", "какая", "какой", "какие", "какую", "пожалуйста", "керек", "bormi", "есть", "нужен", "нужна",
    "покажи", "показать", "ayting", "курсатинг", "ko'rsating", "korsating", "нарх", "нархи", "канча",
    "здравствуйте", "здравствуй", "привет", "добрый", "день", "вечер", "утро",
    "assalomu", "alaykum", "salom", "iltimos", "rahmat", "рахмат", "айтинг",
    "кредит", "кредита", "рассрочка", "рассрочку", "рассрочке", "ежемесячно",
    "nasiya", "nasiyaga", "kredit", "bolib", "bo'lib", "tolash", "to'lash", "muddatli", "oyma", "oy",
    "беру", "оформите", "оформить", "заказываю", "заказ", "отправьте", "доставьте",
    "olaman", "buyurtma", "yuboring", "yetkazib", "bering", "можно", "хочу", "хотел", "хотела",
    "память", "цвет", "xotira", "rang", "модель", "model",
    "и", "или", "а", "на", "в", "за", "для", "без", "первоначального", "взноса", "месяц", "частями",
    "va", "yoki", "ham", "uchun", "bilan", "boshlang'ich", "boshlangich", "tulovsiz", "to'lovsiz", "qilib",
    "я", "мы", "мне", "можно", "ли", "купить", "оплатить", "платить", "полностью", "частями",
    "men", "biz", "sotib", "olish", "mumkinmi", "tolash", "to'lash",
    "спасибо", "ок", "хорошо", "понял", "поняла", "tushunarli", "hop", "xo'p",
    "утром", "днем", "вечером", "завтра", "сегодня", "после", "до", "ertalab", "kunduzi", "kechqurun", "ertaga", "bugun", "soat",
    "удобно", "удобный", "удобное", "время", "qulay", "vaqt",
    "қанча", "нарх", "нархи", "нархини", "айтинг", "керак", "борми", "илтимос", "модел",
    "бўлиб", "тўлаш", "тўлов", "насия", "ойма",
    "наличие", "наличии", "имеется", "mavjud", "mavjudligi", "mavjudmi", "мавжуд", "мавжудлиги", "мавжудми", "бор",
    "доставка", "доставку", "доставкой", "доставить", "привезти", "yetkazish", "berish", "етказиб", "бериш",
    "самовывоз", "самовывозом", "забрать", "olib", "ketish", "олиб", "кетиш",
}
KNOWN_COLORS = {
    "black", "white", "blue", "green", "red", "pink", "gold", "silver", "gray", "grey",
    "natural", "desert", "navy", "purple", "violet", "orange", "yellow", "beige", "brown",
    "space", "midnight", "starlight", "graphite", "cream", "mint", "teal",
    "titanium", "titan", "титан",
    "mocha", "ultramarine",
}

UNSAFE_FUZZY_WORDS = {"pro", "max", "ultra", "plus", "mini", "air", "note", "phone", "watch", "pad"}
BARE_MEMORY_VALUES = {"32", "64", "128", "256", "512", "1024", "2048"}
MEMORY_CONTEXT_BRANDS = {
    "apple", "iphone", "ipad", "macbook", "samsung", "galaxy", "xiaomi",
    "redmi", "poco", "honor", "huawei", "oppo", "vivo", "realme", "oneplus",
    "tecno", "infinix", "google", "pixel",
}
_INLINE_UZ_PHONE = re.compile(
    r"(?<!\d)\+998[\s().\-/]*\d{2}[\s().\-/]*\d{3}"
    r"[\s().\-/]*\d{2}[\s().\-/]*\d{2}(?!\d)"
)


def normalize_model(value: str) -> str:
    value = str(value).casefold().replace("ё", "е").replace("’", "'").replace("‘", "'").replace("ʻ", "'")
    value = re.sub(r"(?<=[a-zа-яўқғҳ0-9])\+", " plus ", value)
    value = re.sub(r"(?<=\d)(?=(?:pro|max|plus|ultra|про|макс|плюс|ультра)\b)", " ", value)
    attached_brands = {
        "айфон": "iphone", "ayfon": "iphone", "aifon": "iphone", "iphone": "iphone",
        "самсунг": "samsung", "samsung": "samsung", "редми": "redmi", "redmi": "redmi",
        "поко": "poco", "poco": "poco", "ipad": "ipad", "airpods": "airpods",
    }
    for old, new in attached_brands.items():
        value = re.sub(rf"(?<![a-zа-яўқғҳ0-9]){re.escape(old)}(?=\d)", new + " ", value)
    for old, new in ALIASES.items():
        value = re.sub(rf"\b{re.escape(old)}\b", new, value)
    value = re.sub(r"\b(\d+)(?:st|nd|rd|th)\b", r"\1", value)
    value = re.sub(r"\b(?:second|second-generation)\s+generation\b", "2", value)
    value = re.sub(r"(\d+(?:\.\d+)?)\s*(гб|gb)\b", r"\1 gb", value)
    value = re.sub(r"(\d+(?:\.\d+)?)\s*(тб|tb)\b", r"\1 tb", value)
    tokens = re.findall(r"[a-zа-яўқғҳ0-9]+", value)
    lookalikes = str.maketrans({"с": "s", "х": "x", "а": "a", "м": "m", "в": "b", "е": "e", "к": "k"})
    tokens = [token.translate(lookalikes) if any(char.isdigit() for char in token) else token for token in tokens]
    return " ".join(tokens)


def extract_product_query(value: str) -> tuple[str, str | None, str | None]:
    """Remove conversational filler and optional memory/color filters."""
    value = _INLINE_UZ_PHONE.sub(" ", str(value or ""))
    raw = value.casefold().replace("’", "'").replace("‘", "'").replace("ʻ", "'")
    slash_memory = re.search(r"\b(\d{1,2})\s*/\s*(\d{2,4})\b", raw)
    normalized = normalize_model(value)
    memory_match = re.search(r"\b(\d+(?:\.\d+)?)\s*(gb|tb)\b", normalized)
    # The approved catalogue uses its ``memory`` column for watch dimensions.
    # Treat an explicit mm value as the selectable attribute while keeping the
    # public return shape backward compatible (query, attribute, color).
    size_match = re.search(r"\b(\d+(?:\.\d+)?)\s*mm\b", normalized)
    words_before_filter = normalized.split()
    bare_candidates = [word for word in words_before_filter if word in BARE_MEMORY_VALUES]
    non_filter_words = [
        word
        for word in words_before_filter
        if word not in BARE_MEMORY_VALUES
        and word not in KNOWN_COLORS
        and word not in SEARCH_FILLERS
    ]
    has_model_number = any(
        candidate != word
        and any(char.isalpha() for char in word)
        and any(char.isdigit() for char in word)
        for candidate in bare_candidates
        for word in words_before_filter
    )
    has_numbered_model_phrase = (
        len(bare_candidates) >= 2
        or any(word in {"pro", "max", "plus", "ultra", "mini", "note"} for word in words_before_filter)
        and any(word.isdigit() and word not in BARE_MEMORY_VALUES for word in words_before_filter)
    )
    bare_memory = (
        bare_candidates[-1]
        if bare_candidates and (
            len(words_before_filter) == 1
            or not non_filter_words
            or bool(MEMORY_CONTEXT_BRANDS.intersection(words_before_filter))
            or has_model_number
            or has_numbered_model_phrase
        )
        else None
    )
    memory = (
        " ".join(memory_match.groups()).upper()
        if memory_match
        else f"{size_match.group(1)}mm"
        if size_match
        else f"{slash_memory.group(2)} GB"
        if slash_memory
        else f"{bare_memory} GB"
        if bare_memory
        else None
    )
    if memory_match:
        normalized = (normalized[: memory_match.start()] + " " + normalized[memory_match.end() :]).strip()
    elif size_match:
        normalized = (
            normalized[: size_match.start()] + " " + normalized[size_match.end() :]
        ).strip()
    elif slash_memory:
        normalized = re.sub(rf"\b{re.escape(slash_memory.group(1))}\s+{re.escape(slash_memory.group(2))}\b", " ", normalized, count=1)
    elif bare_memory:
        normalized = re.sub(rf"\b{re.escape(bare_memory)}\b", " ", normalized, count=1)
    words = normalized.split()
    color_words = [word for word in words if word in KNOWN_COLORS]
    explicit_color_words: list[str] = []
    for index, word in enumerate(words):
        if word not in {"цвет", "rang", "ранг"}:
            continue
        for candidate in words[index + 1 : index + 4]:
            if (
                candidate in SEARCH_FILLERS
                or candidate in BARE_MEMORY_VALUES
                or candidate in {"gb", "tb", "память", "xotira", "хотира", "model", "модель"}
                or any(char.isdigit() for char in candidate)
            ):
                break
            explicit_color_words.append(candidate)
        break
    selected_color_words = explicit_color_words or color_words
    color = " ".join(selected_color_words) or None
    removed_color_words = set(color_words) | set(explicit_color_words)
    query_words = [
        word
        for word in words
        if word not in SEARCH_FILLERS and word not in removed_color_words
    ]
    return " ".join(query_words), memory, color


def safe_product_url(value: object) -> str | None:
    url = str(value or "").strip()
    if len(url) > 500 or any(char.isspace() or ord(char) < 32 for char in url):
        return None
    try:
        parsed = urlparse(url)
        port = parsed.port
    except ValueError:
        return None
    if (
        parsed.scheme != "https"
        or parsed.hostname not in TRUSTED_POST_HOSTS
        or parsed.username is not None
        or parsed.password is not None
        or port is not None
    ):
        return None
    parts = [part for part in parsed.path.split("/") if part]
    if len(parts) < 2 or not parts[-1].isdigit():
        return None
    return url


def _product_id(value: object) -> int | None:
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    parsed = int(text) if text.isdigit() else None
    return parsed if parsed and parsed > 0 else None


def _model_alias_keys(value: str) -> tuple[str, ...]:
    normalized = normalize_model(_family_name(value))
    if not normalized:
        return ()
    keys = [normalized]
    tokens = normalized.split()
    # The price source sometimes omits the manufacturer prefix used in
    # Bot_URLS. Add only conservative aliases and keep them only if unique.
    if len(tokens) >= 3 and tokens[0] in {"apple", "xiaomi"}:
        keys.append(" ".join(tokens[1:]))
    if len(tokens) >= 4 and tokens[:2] == ["samsung", "galaxy"]:
        keys.append(" ".join([tokens[0], *tokens[2:]]))
    return tuple(dict.fromkeys(keys))


def _load_product_url_mappings(
    path: Path | str | None = None,
) -> tuple[dict[int, str], dict[str, str], frozenset[str]]:
    """Load ID links and only unambiguous model-name fallbacks."""
    project_root = Path(__file__).resolve().parents[1]
    configured = os.getenv("PRODUCT_URLS_PATH", "").strip()
    candidates = [
        Path(path) if path else None,
        Path(configured) if configured else None,
        project_root / "data" / "Bot_URLS.xlsx",
        project_root / "Data" / "Bot_URLS.xlsx",
        project_root / "data" / "Data" / "Bot_URLS.xlsx",
        project_root / "Bot_URLS.xlsx",
    ]
    source = next((candidate for candidate in candidates if candidate and candidate.is_file()), None)
    if source is None:
        return {}, {}, frozenset()

    from openpyxl import load_workbook

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip().lstrip("\ufeff").casefold() for value in next(rows, ())]
        try:
            id_index = headers.index("product_id")
            url_index = headers.index("post_id")
        except ValueError:
            return {}, {}, frozenset()
        model_index = headers.index("model") if "model" in headers else None

        result: dict[int, str] = {}
        model_candidates: dict[str, set[str]] = {}
        model_keys: set[str] = set()
        for row in rows:
            raw_ids = row[id_index] if id_index < len(row) else None
            raw_url = row[url_index] if url_index < len(row) else None
            url = safe_product_url(raw_url)
            if not url:
                continue
            for raw_id in str(raw_ids or "").split(","):
                product_id = _product_id(raw_id)
                if product_id is not None:
                    # Match Seller_Bot.py: the first post for a product wins.
                    result.setdefault(product_id, url)
            if model_index is not None and model_index < len(row):
                model_name = str(row[model_index] or "").strip()
                for model_key in _model_alias_keys(model_name):
                    model_keys.add(model_key)
                    model_candidates.setdefault(model_key, set()).add(url)
        # Never guess between two posts with the same normalized model title.
        unique_models = {
            model: next(iter(urls))
            for model, urls in model_candidates.items()
            if len(urls) == 1
        }
        return result, unique_models, frozenset(model_keys)
    finally:
        workbook.close()


def load_product_urls(path: Path | str | None = None) -> dict[int, str]:
    """Load product_id -> Telegram post exactly like Seller_Bot.py."""
    return _load_product_url_mappings(path)[0]


def load_model_urls(path: Path | str | None = None) -> dict[str, str]:
    """Load unique normalized model -> Telegram post fallback mapping."""
    return _load_product_url_mappings(path)[1]


def load_product_model_keys(path: Path | str | None = None) -> frozenset[str]:
    """Return normalized catalogue names without loading prices or using network."""
    return _load_product_url_mappings(path)[2]


def _model_url(variants: list[ProductVariant] | tuple[ProductVariant, ...]) -> str | None:
    return next((variant.url for variant in variants if variant.url), None)


def _family_name(value: str) -> str:
    name = str(value or "").strip()
    patterns = (
        r"\s*\(\s*(?:e?sim)(?:\s*\+\s*e?sim)?\s*\)\s*$",
        r"\s*\(\s*(?:global\s+(?:rom|version)|china\s+version|hong\s+kong|cn|eu|usa|uae|hk)\s*\)\s*$",
        r"\s+(?:[345]g|usb(?:[\s-]*type)?[\s-]*c|lightning)\s*$",
    )
    # Some rows contain more than one technical suffix, e.g. "(eSIM) 5G".
    # Re-run the safe suffix list until it is stable.
    for _ in range(3):
        previous = name
        for pattern in patterns:
            name = re.sub(pattern, "", name, flags=re.I).strip()
        if name == previous:
            break
    return name.strip()


def _fuzzy_token_score(query_words: list[str], model_words: list[str]) -> float | None:
    if not query_words or (len(query_words) == 1 and query_words[0] in UNSAFE_FUZZY_WORDS):
        return None
    scores: list[float] = []
    for query_word in query_words:
        if query_word in model_words:
            scores.append(1.0)
            continue
        # Numbers and very short tokens identify product generations and must
        # never be silently changed by fuzzy matching.
        if any(char.isdigit() for char in query_word) or len(query_word) <= 3:
            return None
        score = max((difflib.SequenceMatcher(None, query_word, word).ratio() for word in model_words), default=0.0)
        if score < 0.74:
            return None
        scores.append(score)
    average = sum(scores) / len(scores)
    return average if average >= 0.86 else None


def _memory_gb(value: str | None) -> Decimal | None:
    normalized = normalize_model(value or "")
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*(gb|tb)", normalized)
    if not matches:
        return None
    number, unit = matches[-1]
    amount = Decimal(number)
    return amount * Decimal(1024) if unit == "tb" else amount


def _memory_matches(requested: str, available: str) -> bool:
    requested_gb, available_gb = _memory_gb(requested), _memory_gb(available)
    if requested_gb is not None and available_gb is not None:
        return requested_gb == available_gb
    return normalize_model(requested) in normalize_model(available)


class ExistingGoogleProductRepository:
    """Read-only adapter for the project's approved bot_prices Google sheet."""

    def __init__(self, max_age_minutes: int = 1440, urls_path: Path | str | None = None):
        self.max_age = max_age_minutes * 60
        self.urls_path = urls_path
        self._loaded = 0.0
        self._variants: list[ProductVariant] = []
        self._known_model_keys: frozenset[str] | None = None
        self._lock = threading.Lock()

    def recognizes_query(self, query: str) -> bool:
        """Recognize an exact approved card title without fetching price data."""
        if self._known_model_keys is None:
            self._known_model_keys = load_product_model_keys(self.urls_path)
        normalized = normalize_model(query)
        return bool(normalized and normalized in self._known_model_keys)

    def _load(self) -> None:
        if self._variants and time.time() - self._loaded < min(self.max_age, 300):
            return
        with self._lock:
            if self._variants and time.time() - self._loaded < min(self.max_age, 300):
                return
            from instagram_bot import PRODUCTS_CACHE, get_product_catalog

            catalog = get_product_catalog(force_refresh=True)
            # The existing integration deliberately returns stale cache after a
            # failed Google request. Business replies must not present that as a
            # current approved price.
            if PRODUCTS_CACHE.get("source") != "google_sheets":
                raise RuntimeError("approved product source is unavailable or stale")
            rows = catalog.get("rows", []) if isinstance(catalog, dict) else []
            product_urls, model_urls, model_keys = _load_product_url_mappings(self.urls_path)
            self._known_model_keys = model_keys
            rate = Decimal(str(catalog.get("kurs", 1))) if isinstance(catalog, dict) else Decimal(1)
            if not rate.is_finite() or rate <= 0:
                raise RuntimeError("approved product source returned an invalid exchange rate")
            parsed: list[ProductVariant] = []
            for row in rows:
                model = str(row.get("model_name", "")).strip()
                family_name = str(row.get("family_name") or _family_name(model)).strip() or model
                price = row.get("price")
                product_id = _product_id(row.get("product_id"))
                if not model or price is None or price == "":
                    continue
                converted_price = Decimal(str(price)) * rate
                if not converted_price.is_finite() or converted_price <= 0:
                    continue
                warranty = row.get("warranty_period")
                try:
                    warranty_months = int(warranty) if warranty is not None else None
                except (TypeError, ValueError):
                    warranty_months = None
                parsed.append(
                    ProductVariant(
                        model=model,
                        memory=str(row.get("memory", "")).strip(),
                        color=str(row.get("color", "")).strip(),
                        price_uzs=converted_price,
                        product_id=product_id,
                        url=(
                            product_urls.get(product_id)
                            if product_id is not None and product_urls.get(product_id)
                            else next(
                                (model_urls[key] for key in _model_alias_keys(family_name) if key in model_urls),
                                None,
                            )
                        ),
                        warranty_months=warranty_months,
                    )
                )
            if not parsed:
                raise RuntimeError("approved product source returned no rows")
            self._variants = parsed
            self._loaded = time.time()

    def search(self, query: str, memory: str | None = None, color: str | None = None) -> ProductMatch:
        self._load()
        if self.max_age <= 0 or time.time() - self._loaded > self.max_age:
            raise RuntimeError("product prices are stale")

        extracted_query, extracted_memory, extracted_color = extract_product_query(query)
        normalized_query = normalize_model(extracted_query)
        memory = memory or extracted_memory
        color = color or extracted_color
        if not normalized_query:
            return ProductMatch("not_found")
        query_words = normalized_query.split()

        def found_match(
            selected: str,
            unfiltered: list[ProductVariant],
        ) -> ProductMatch:
            memory_matches = (
                [
                    variant for variant in unfiltered
                    if _memory_matches(memory, variant.memory)
                ]
                if memory else unfiltered
            )
            color_matches = (
                [
                    variant for variant in unfiltered
                    if normalize_model(color) in normalize_model(variant.color)
                ]
                if color else unfiltered
            )
            variants = [
                variant for variant in unfiltered
                if (not memory or variant in memory_matches)
                and (not color or variant in color_matches)
            ]
            unmatched: list[str] = []
            if memory and not memory_matches:
                unmatched.append("memory")
            if color and not color_matches:
                unmatched.append("color")
            if not variants and not unmatched and memory and color:
                unmatched.append("memory_color_combination")
            if not variants:
                if memory and memory_matches:
                    variants = memory_matches
                elif color and color_matches:
                    variants = color_matches
                else:
                    variants = unfiltered
            selected_url = _model_url(unfiltered)
            return ProductMatch(
                "found",
                (selected,),
                tuple(variants),
                ((selected, selected_url),) if selected_url else (),
                requested_memory=memory,
                requested_color=color,
                unmatched_filters=tuple(unmatched),
                all_variants=tuple(unfiltered),
            )

        raw_exact = [
            variant for variant in self._variants
            if normalize_model(variant.model) == normalized_query
        ]
        if raw_exact:
            return found_match(raw_exact[0].model, raw_exact)

        by_model: dict[str, list[ProductVariant]] = {}
        for variant in self._variants:
            family = _family_name(variant.model) or variant.model
            by_model.setdefault(family, []).append(variant)

        if len(query_words) == 1 and query_words[0] in UNSAFE_FUZZY_WORDS:
            return ProductMatch("not_found")

        ranked: list[tuple[int, float, int, str]] = []
        for model in by_model:
            normalized = normalize_model(model)
            family = normalize_model(_family_name(model))
            searchable = family or normalized
            searchable_words = searchable.split()
            # A fuzzy whole-string score must never silently substitute a
            # product generation (16 -> 15, S26 -> S25, X8 -> X7).
            numeric_query_words = [word for word in query_words if any(char.isdigit() for char in word)]
            if any(word not in searchable_words for word in numeric_query_words):
                continue
            if normalized == normalized_query:
                rank, ratio = 0, 1.0
            elif family == normalized_query or normalized_query in _model_alias_keys(model):
                rank, ratio = 1, 1.0
            elif searchable.startswith(normalized_query):
                rank, ratio = 2, 1.0
            elif normalized_query in searchable:
                rank, ratio = 3, 1.0
            elif all(word in searchable_words for word in query_words):
                rank, ratio = 4, 1.0
            else:
                token_ratio = _fuzzy_token_score(query_words, searchable_words)
                whole_ratio = difflib.SequenceMatcher(None, normalized_query, searchable).ratio()
                ratio = max(token_ratio or 0.0, whole_ratio if len(query_words) >= 2 else 0.0)
                if ratio < 0.86:
                    continue
                rank = 5
            ranked.append((rank, -ratio, len(normalized), model))
        ranked.sort(key=lambda item: (item[0], item[1], item[2], item[3].casefold()))
        candidates = [item[3] for item in ranked[:5]]
        if not candidates:
            return ProductMatch("not_found")

        model_urls = tuple(
            (model, url)
            for model in candidates
            if (url := _model_url(by_model[model]))
        )
        exact = [model for model in candidates if normalize_model(model) == normalized_query]
        family_exact = [model for model in candidates if normalize_model(_family_name(model)) == normalized_query]
        alias_exact = [model for model in candidates if normalized_query in _model_alias_keys(model)]
        if len(exact) == 1:
            candidates = exact
        elif len(family_exact) == 1:
            candidates = family_exact
        elif len(alias_exact) == 1:
            candidates = alias_exact
        elif len(candidates) > 1:
            return ProductMatch("ambiguous", tuple(candidates), (), model_urls)

        selected = candidates[0]
        return found_match(selected, list(by_model[selected]))


def _format_amount(value: Decimal, language: str) -> str:
    quantum = Decimal("1000") if value > Decimal("10000") else Decimal("1")
    rounded = int((value / quantum).quantize(Decimal("1"), rounding=ROUND_HALF_UP) * quantum)
    amount = f"{rounded:,}".replace(",", " ")
    # Keep one familiar currency spelling in every localized screen.  Mixing
    # ``сум`` and ``so‘m`` made the same catalogue row look like two prices.
    return f"{amount} so'm"


def _memory_sort(value: str) -> tuple[int, str]:
    normalized = normalize_model(value)
    capacity = _memory_gb(value)
    return (int(capacity * 100) if capacity is not None else 10**9, normalized)


def _bounded_html_message(title: str, lines: list[str], tail: str, limit: int = 4096) -> str:
    """Fit complete escaped lines; never slice through an HTML tag/entity."""
    # Callers already bound raw model/URL fields, so truncating HTML itself is
    # unnecessary and could split a tag or entity.
    tail = tail[:1200]
    kept: list[str] = []
    suffix = "\n\n" + tail
    for line in lines:
        candidate = title + "\n\n" + "\n".join([*kept, line]) + suffix
        if len(candidate.encode("utf-16-le")) // 2 > limit:
            break
        kept.append(line)
    if not kept and lines:
        # Price lines are bounded by source column sizes in practice. This
        # defensive fallback omits an oversized line instead of breaking HTML.
        fallback = "Варианты есть в базе." if "Цены" in title else "Variantlar bazada mavjud."
        kept.append(fallback)
    message = title + ("\n\n" + "\n".join(kept) if kept else "") + suffix
    return message if len(message.encode("utf-16-le")) // 2 <= limit else title


def _filter_mismatch_notice(match: ProductMatch, language: str) -> str | None:
    if not match.unmatched_filters:
        return None
    memory = html.escape(str(match.requested_memory or "")[:80])
    color = html.escape(str(match.requested_color or "")[:80])
    if "memory_color_combination" in match.unmatched_filters:
        detail_ru = f"сочетанию памяти {memory} и цвета {color}"
        detail_uz = f"{memory} xotira va {color} rang kombinatsiyasiga"
    else:
        ru_parts = []
        uz_parts = []
        if "memory" in match.unmatched_filters:
            ru_parts.append(f"памяти {memory}")
            uz_parts.append(f"{memory} xotiraga")
        if "color" in match.unmatched_filters:
            ru_parts.append(f"цвету {color}")
            uz_parts.append(f"{color} rangga")
        detail_ru = " и ".join(ru_parts)
        detail_uz = " va ".join(uz_parts)
    ru = (
        f"В текущей базе цен нет точного варианта по {detail_ru}. "
        "Ниже показаны другие варианты модели; наличие подтвердит менеджер."
    )
    uz = (
        f"Joriy narxlar bazasida {detail_uz} mos aniq variant yo‘q. "
        "Quyida modelning boshqa variantlari ko‘rsatilgan; mavjudligini menejer tasdiqlaydi."
    )
    return f"{ru} / {uz}" if language == "bi" else ru if language == "ru" else uz


def format_result(match: ProductMatch, language: str) -> str:
    model = match.models[0]
    model_url = safe_product_url(match.url_for(model) or _model_url(match.variants))
    safe_model = html.escape(model[:300])
    if model_url:
        safe_model = f'<a href="{html.escape(model_url, quote=True)}">{safe_model}</a>'

    notice = _filter_mismatch_notice(match, language)
    if len(match.variants) > 12:
        grouped: dict[str, Decimal] = {}
        for variant in match.variants:
            key = variant.memory or ("Вариант" if language == "ru" else "Variant")
            grouped[key] = min(grouped.get(key, variant.price_uzs), variant.price_uzs)
        if language == "ru":
            lines = [
                f"{html.escape(str(memory)[:160])} — от {_format_amount(price, language)}"
                for memory, price in sorted(grouped.items(), key=lambda item: _memory_sort(item[0]))
            ]
        elif language == "uz":
            lines = [
                f"{html.escape(str(memory)[:160])} — {_format_amount(price, language)} dan boshlab"
                for memory, price in sorted(grouped.items(), key=lambda item: _memory_sort(item[0]))
            ]
        else:
            lines = [
                f"{html.escape(str(memory)[:160])} — от / {_format_amount(price, language)} dan boshlab"
                for memory, price in sorted(grouped.items(), key=lambda item: _memory_sort(item[0]))
            ]
    else:
        grouped: dict[tuple[str, Decimal], list[str]] = {}
        for variant in sorted(match.variants, key=lambda item: (_memory_sort(item.memory), item.price_uzs, item.color)):
            key = (variant.memory, variant.price_uzs)
            colors = grouped.setdefault(key, [])
            if variant.color and variant.color not in colors:
                colors.append(variant.color)
        lines = []
        for (memory, price), colors in grouped.items():
            parts = [html.escape(str(value)[:600]) for value in (memory, ", ".join(colors)) if value]
            label = ", ".join(parts) or ("Вариант" if language == "ru" else "Variant")
            lines.append(f"{label} — {_format_amount(price, language)}")

    if notice:
        lines.insert(0, notice)

    if language == "bi":
        title = safe_model
        tail = "Выберите вариант. / Variantni tanlang."
    elif language == "ru":
        title = safe_model
        tail = "Выберите вариант."
    else:
        title = safe_model
        tail = "Variantni tanlang."
    return _bounded_html_message(title, lines, tail)


def format_ambiguous_result(match: ProductMatch, language: str) -> str:
    if language == "bi":
        intro = "Выберите модель / Modelni tanlang:"
        tail = "Нажмите кнопку ниже. / Quyidagi tugmani bosing."
    else:
        intro = "Выберите модель:" if language == "ru" else "Modelni tanlang:"
        tail = "Нажмите кнопку ниже." if language == "ru" else "Quyidagi tugmani bosing."
    lines = []
    for number, model in enumerate(match.models[:5], 1):
        safe_model = html.escape(model[:300])
        url = safe_product_url(match.url_for(model))
        if url:
            safe_model = f'<a href="{html.escape(url, quote=True)}">{safe_model}</a>'
        lines.append(f"{number}. {safe_model}")
    return _bounded_html_message(intro, lines, tail)


def model_link_keyboard(match: ProductMatch) -> dict | None:
    rows = []
    for number, model in enumerate(match.models[:5], 1):
        if url := safe_product_url(match.url_for(model)):
            label = " ".join(str(model).split())
            rows.append([{"text": f"{number}. {label}"[:64], "url": url}])
    return {"inline_keyboard": rows} if rows else None
